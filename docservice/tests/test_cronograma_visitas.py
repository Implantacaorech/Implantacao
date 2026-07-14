# -*- coding: utf-8 -*-
"""Testes do endpoint /gerar/cronograma-visitas — o único endpoint desta fatia (item 3
da migração). Verifica que o .xlsx real sai com cabeçalho, linhas de visita e analista
responsável, replicando webapp/test_painel.py:test_cronograma_xlsx_cabecalho_e_linhas."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import io

from main import app

client = TestClient(app)


def payload_base():
    return {
        "projeto": {
            "id": 1,
            "cliente": "Cliente Teste LTDA",
            "cnpj": "00.000.000/0001-00",
            "numeroProjeto": "PRJ-1",
            "consultor": "Ana Consultora",
            "horasCobradas": "40",
            "horasBonificadas": "8",
        },
        "atividades": [
            {
                "id": 1,
                "modulo": "FAT",
                "seq": 1,
                "descricao": "Cadastro de produtos",
                "tipo": "Treinamento",
                "data": "2026-08-10",
                "turno": "manha",
                "tecnico": "Ana Consultora",
                "status": "Agendada",
            },
            {
                "id": 2,
                "modulo": "EST",
                "seq": 1,
                "descricao": "Inventário",
                "tipo": "Treinamento",
                "data": "2026-08-11",
                "turno": "tarde",
                "tecnico": "Beto",
                "status": "Solicitada",  # sem data/turno seria ignorado; aqui tem os dois
            },
        ],
        "horarios": {"manha": {"inicio": "08:00", "fim": "12:00"}, "tarde": {"inicio": "13:00", "fim": "17:00"}},
        "designacoes": [
            {"modulo": "FAT", "consultor": "Ana Consultora", "analista": "Carla"},
            {"modulo": "EST", "consultor": "Beto", "analista": ""},
        ],
        "cronogramaConfig": {"analistaPadrao": "Diego"},
    }


def test_health():
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json() == {"status": "ok"}


def test_gera_xlsx_com_cabecalho_e_linhas():
    r = client.post("/gerar/cronograma-visitas", json=payload_base())
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    # Checagem estrita do travessão (U+2014, não um "?"/replacement char) — achado real desta
    # migração: sem PYTHONUTF8=1, este literal (vindo de gl_xlsx.py, copiado de webapp/) saía
    # corrompido pela codepage do Windows. Ver a checagem em main.py e docservice/iniciar.bat.
    assert ws["C2"].value == "Cronograma de Visitas — Cliente Teste LTDA"
    assert ord(ws["C2"].value[22]) == 0x2014
    assert ws["D3"].value == "Cliente Teste LTDA"
    assert ws["G3"].value == "00.000.000/0001-00"
    assert ws["D4"].value == "PRJ-1"
    assert ws["G4"].value == "Ana Consultora"

    cabecalho = [c.value for c in ws[6]]  # linha 6 = cabeçalho da tabela (linhas 1-5 são o topo)
    assert "Analista Responsável" in cabecalho
    assert "Módulo" in cabecalho and "Técnico" in cabecalho

    linhas = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=7, max_row=8)]
    # FAT (analista da designação = Carla) vem antes de EST (sem override -> analista padrão Diego)
    assert linhas[0][4] == "FAT"
    assert linhas[1][6] == "Inventário"  # descrição acentuada — outra checagem de encoding
    assert linhas[0][9] == "Carla"
    assert linhas[1][4] == "EST"
    assert linhas[1][9] == "Diego"


def test_rejeita_quando_nenhuma_atividade_alocada():
    payload = payload_base()
    for a in payload["atividades"]:
        a["data"], a["turno"] = "", ""
    r = client.post("/gerar/cronograma-visitas", json=payload)
    assert r.status_code == 422
    assert "Nenhuma atividade alocada" in r.json()["detail"]


def test_designacoes_e_config_sao_isoladas_por_requisicao():
    """Duas chamadas seguidas com designações diferentes não podem vazar contexto uma
    para a outra — o contextvar precisa ser sobrescrito a cada requisição."""
    p1 = payload_base()
    p1["designacoes"] = [{"modulo": "FAT", "consultor": "Ana Consultora", "analista": "Analista-1"}]
    r1 = client.post("/gerar/cronograma-visitas", json=p1)
    wb1 = load_workbook(io.BytesIO(r1.content))
    linha_fat_1 = next(row for row in wb1.active.iter_rows(min_row=7, max_row=8) if row[4].value == "FAT")
    assert linha_fat_1[9].value == "Analista-1"

    p2 = payload_base()
    p2["designacoes"] = [{"modulo": "FAT", "consultor": "Ana Consultora", "analista": "Analista-2"}]
    r2 = client.post("/gerar/cronograma-visitas", json=p2)
    wb2 = load_workbook(io.BytesIO(r2.content))
    linha_fat_2 = next(row for row in wb2.active.iter_rows(min_row=7, max_row=8) if row[4].value == "FAT")
    assert linha_fat_2[9].value == "Analista-2"
