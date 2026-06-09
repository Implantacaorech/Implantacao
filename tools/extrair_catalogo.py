# -*- coding: utf-8 -*-
"""
extrair_catalogo.py — gera tools/data/catalogo_modulos.yaml a partir da planilha
"Roteiro e Check List implantação por módulos.xlsx" (aba "Lista de Módulos").

Catálogo = código (col A) + abreviação (col B) + descrição (col C) + **área do
Levantamento** (derivada do bloco/posição na lista; EDITÁVEL no YAML).

Reexecute quando a planilha mudar:
    python extrair_catalogo.py ["caminho/Roteiro e Check List ....xlsx"]
"""
import os
import sys
import glob
import yaml
from openpyxl import load_workbook

import _common as C

# Abreviação que DEFINE a área do bloco (ao percorrer a Lista de cima para baixo).
HEADER_AREA = {
    "CTB": "Gestão Fiscal, Contábil e Patrimonial",
    "GPA": "Gestão Fiscal, Contábil e Patrimonial",
    "LFI": "Gestão Fiscal, Contábil e Patrimonial",
    "FPA": "Folha de Pagamento",
    "FAT": "Vendas e Faturamento",
    "TLO": "Vendas e Faturamento",
    "GCA": "Vendas e Faturamento",
    "PDV": "Vendas e Faturamento",
    "OSE": "Vendas e Faturamento",
    "TEL": "Vendas e Faturamento",
    "FIN": "Gestão Financeira",
    "AUE": "Gestão Financeira",
    "EST": "Compras/Estoque",
    "COM": "Compras/Estoque",
    "WMS": "Compras/Estoque",
    "GIN": "Produção",
    "MHB": "BI e Integrações",
    "AWR": "BI e Integrações",
}
# Override por abreviação (tem prioridade sobre a área do bloco).
OVERRIDE = {
    "FSE": "Folha de Pagamento", "CMO": "Folha de Pagamento", "CEF": "Folha de Pagamento",
    "RHU": "Recursos Humanos",
    "SER": "Recrutamento e Seleção", "RSE": "Recrutamento e Seleção",
    "TRN": "Treinamentos", "SAO": "Saúde Ocupacional",
    "STR": "Segurança do Trabalho", "STT": "Segurança do Trabalho",
    "CSA": "Cargos e Salários", "CSG": "Cargos e Salários",
    "AVF": "Avaliação e Feedback",
    "PGP": "Portal de Funcionários", "PWC": "Portal de Funcionários",
    "PVA": "Portal de Vagas",
    "CEE": "Comércio Exterior", "CEI": "Comércio Exterior",
    "CCR": "BI e Integrações", "RME": "BI e Integrações", "GVT": "BI e Integrações",
    "EDU": "BI e Integrações",
}


def main(path=None):
    if not path:
        cands = glob.glob(os.path.expanduser("~/Downloads/Roteiro e Check List*.xlsx"))
        if not cands:
            sys.exit("Informe o caminho da planilha 'Roteiro e Check List ....xlsx'.")
        path = cands[0]
    wb = load_workbook(path, data_only=True)
    ws = wb["Lista de Módulos"]

    modulos, atual = [], "Outros"
    for r in range(2, ws.max_row + 1):
        cod = ws.cell(r, 1).value
        ab = ws.cell(r, 2).value
        desc = ws.cell(r, 3).value
        ab = str(ab).strip() if ab else ""
        if not ab:
            continue
        if ab in HEADER_AREA:
            atual = HEADER_AREA[ab]
        area = OVERRIDE.get(ab, atual)
        modulos.append({
            "codigo": int(cod) if isinstance(cod, (int, float)) else (str(cod).strip() if cod else ""),
            "abrev": ab,
            "descricao": str(desc).strip() if desc else "",
            "area": area,
        })

    out = os.path.join(C.DATA, "catalogo_modulos.yaml")
    with open(out, "w", encoding="utf-8") as f:
        f.write("# Catálogo de módulos/adicionais do SIGER (extraído do Roteiro e Check List).\n")
        f.write("# 'area' = área do Levantamento. EDITÁVEL: ajuste o agrupamento se necessário.\n")
        yaml.safe_dump({"modulos": modulos}, f, allow_unicode=True, sort_keys=False, width=120)

    # resumo por área
    from collections import Counter
    cont = Counter(m["area"] for m in modulos)
    print("OK: %d módulos -> %s" % (len(modulos), out))
    for area, n in cont.most_common():
        print("   %3d  %s" % (n, area))


if __name__ == "__main__":
    main(*sys.argv[1:])
