# -*- coding: utf-8 -*-
"""Gera a planilha de Reconciliação de Conversão (.xlsx): confere contagem e
valores origem × destino por entidade, registra as cargas (mock loads) e o
sign-off dos dados convertidos.

Uso:
    python gerar_reconciliacao_conversao.py [data/exemplo_cliente.yaml] [data/conversao.yaml]
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import _common as C

COLS = ["Entidade", "Chave", "RNS COB", "Qtd Origem", "Qtd Destino", "Dif. Qtd",
        "Valor Origem (R$)", "Valor Destino (R$)", "Dif. Valor", "Amostra OK?",
        "Status", "Observação"]
WIDTHS = [34, 16, 12, 12, 12, 10, 16, 16, 12, 12, 14, 40]


def build_capa(wb, cliente, criterios):
    ws = wb.active
    ws.title = "Capa"
    C.set_widths(ws, [26, 60])
    C.title_block(ws, "Reconciliação de Conversão",
                  f"{cliente.get('nome','')} · gerado em {C.today()}", span=2)
    info = [
        ("Cliente", cliente.get("nome", "")),
        ("Código SICLA", cliente.get("codigo_sicla", "")),
        ("Virada prevista", cliente.get("data_virada_prevista", "")),
        ("", ""),
        ("Origem", "Sistema antigo do cliente"),
        ("Destino", "SIGER®"),
    ]
    r = 4
    for k, v in info:
        a = ws.cell(row=r, column=1, value=k)
        if k:
            a.font = C.HEADER_FONT; a.fill = C.HEADER_FILL
        a.alignment = C.WRAP
        ws.cell(row=r, column=2, value=v).alignment = C.WRAP
        r += 1
    ws.cell(row=r + 1, column=1, value="Critérios de aceite").font = C.TITLE_FONT
    for i, crit in enumerate(criterios):
        ws.cell(row=r + 2 + i, column=1, value=f"• {crit}").alignment = C.WRAP
        ws.merge_cells(start_row=r + 2 + i, start_column=1, end_row=r + 2 + i, end_column=2)


def build_reconciliacao(wb, entidades):
    ws = wb.create_sheet("Reconciliação")
    C.header_row(ws, COLS)
    C.set_widths(ws, WIDTHS)
    for i, e in enumerate(entidades):
        r = i + 2
        ws.cell(row=r, column=1, value=e.get("nome", "")).alignment = C.WRAP
        ws.cell(row=r, column=2, value=e.get("chave", "")).alignment = C.WRAP
        ws.cell(row=r, column=3, value=e.get("rns_cob", "")).alignment = C.WRAP
        # D,E vazios (preencher); F = E-D
        ws.cell(row=r, column=6, value=f"=E{r}-D{r}").alignment = C.CENTER
        # G,H vazios; I = H-G
        ws.cell(row=r, column=9, value=f"=H{r}-G{r}").alignment = C.CENTER
        ws.cell(row=r, column=12, value=e.get("obs", "")).alignment = C.WRAP
        for col in (4, 5, 7, 8):
            ws.cell(row=r, column=col).border = C.BORDER
    n = len(entidades)
    dv_st = DataValidation(type="list", formula1='"Conferido,Divergente,Pendente"', allow_blank=True)
    dv_am = DataValidation(type="list", formula1='"Sim,Não,Parcial"', allow_blank=True)
    ws.add_data_validation(dv_st); ws.add_data_validation(dv_am)
    dv_st.add(f"K2:K{n + 1}"); dv_am.add(f"J2:J{n + 1}")


def build_cargas(wb, cargas):
    ws = wb.create_sheet("Cargas (mock loads)")
    cols = ["Carga", "Data", "Resultado", "Divergências encontradas", "Responsável"]
    C.header_row(ws, cols)
    C.set_widths(ws, [22, 16, 18, 44, 20])
    rows = [[c.get("tipo", ""), "", "", "", ""] for c in cargas]
    C.write_rows(ws, rows)
    dv = DataValidation(type="list", formula1='"OK,OK com ressalvas,Refazer"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C2:C{len(rows) + 1}")


def build_signoff(wb):
    ws = wb.create_sheet("Sign-off")
    C.set_widths(ws, [60])
    C.title_block(ws, "Aceite dos dados convertidos", span=1)
    ws.cell(row=4, column=1,
            value="Liberar a conversão oficial somente com contagem e valores conferidos, "
                  "amostra validada e divergências aceitas pelo cliente.").alignment = C.WRAP
    ws.merge_cells("A4:A5")
    for i, papel in enumerate(["Consultor de Implantação", "Equipe de Conversão",
                               "Usuário Líder (cliente)"]):
        r = 8 + i * 2
        ws.cell(row=r, column=1, value="__________________________")
        ws.cell(row=r + 1, column=1, value=papel).font = C.SUB_FONT


def main(cliente_path="data/exemplo_cliente.yaml", conv_path="data/conversao.yaml"):
    cli = C.load_yaml(os.path.basename(cliente_path))
    conv = C.load_yaml(os.path.basename(conv_path))
    cliente = cli.get("cliente", {})

    wb = Workbook()
    build_capa(wb, cliente, conv.get("criterios_aceite", []))
    build_reconciliacao(wb, conv.get("entidades", []))
    build_cargas(wb, conv.get("cargas", []))
    build_signoff(wb)

    C.ensure_out()
    fname = f"Reconciliacao_Conversao_{C.slug(cliente.get('nome'))}.xlsx"
    path = os.path.join(C.OUT, fname)
    wb.save(path)
    print(f"OK: {fname} ({len(conv.get('entidades', []))} entidades) -> {path}")


if __name__ == "__main__":
    args = sys.argv[1:]
    main(*args) if args else main()
