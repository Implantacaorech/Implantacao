# -*- coding: utf-8 -*-
"""Gera o Log de Fit/Gap (.xlsx): registra a aderência de cada processo ao
padrão do SIGER® e a decisão (padrão / configuração / desenvolvimento / fora de
escopo), com resumo de esforço e governança de customização.

Uso:
    python gerar_log_fitgap.py [data/exemplo_cliente.yaml] [data/fitgap.yaml]
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import _common as C

COLS = ["ID", "Processo", "Área", "Aderência", "Decisão", "RNS vinculada",
        "Esforço (h)", "Prioridade", "Status"]
WIDTHS = [8, 44, 18, 20, 20, 16, 12, 12, 16]
SHEET = "Log Fit-Gap"  # já sanitizado (sem "/")


def build_capa(wb, cliente, gov):
    ws = wb.active
    ws.title = "Capa"
    C.set_widths(ws, [26, 64])
    C.title_block(ws, "Log de Fit/Gap — Aderência ao SIGER®",
                  f"{cliente.get('nome','')} · gerado em {C.today()}", span=2)
    info = [("Cliente", cliente.get("nome", "")), ("Código SICLA", cliente.get("codigo_sicla", ""))]
    r = 4
    for k, v in info:
        a = ws.cell(row=r, column=1, value=k); a.font = C.HEADER_FONT; a.fill = C.HEADER_FILL; a.alignment = C.WRAP
        ws.cell(row=r, column=2, value=v).alignment = C.WRAP
        r += 1
    gv = ws.cell(row=r + 1, column=1, value="Governança de customização")
    gv.font = C.TITLE_FONT
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 4, end_column=2)
    ws.cell(row=r + 2, column=1, value=gov).alignment = C.WRAP


def build_log(wb, itens):
    ws = wb.create_sheet(SHEET)
    C.header_row(ws, COLS)
    C.set_widths(ws, WIDTHS)
    rows = []
    for i, it in enumerate(itens, 1):
        rows.append([
            f"FG-{i:02d}", it.get("processo", ""), it.get("area", ""),
            it.get("aderencia", ""), it.get("decisao", ""), it.get("rns", ""),
            it.get("esforco_h", 0), it.get("prioridade", ""), "Em aberto",
        ])
    C.write_rows(ws, rows)
    n = len(rows)
    dv_ad = DataValidation(type="list",
                           formula1='"Standard,Standard (configuração),Parcial,Gap"', allow_blank=True)
    dv_de = DataValidation(type="list",
                           formula1='"Usar padrão,Configuração,Desenvolvimento,Fora de escopo"', allow_blank=True)
    dv_st = DataValidation(type="list", formula1='"Em aberto,Aprovado,Concluído,Cancelado"', allow_blank=True)
    for dv, col in ((dv_ad, "D"), (dv_de, "E"), (dv_st, "I")):
        ws.add_data_validation(dv); dv.add(f"{col}2:{col}{n + 1}")
    return n


def build_resumo(wb, n):
    ws = wb.create_sheet("Resumo")
    C.set_widths(ws, [28, 16, 44])
    C.title_block(ws, "Resumo do Fit/Gap", span=3)
    metr = [
        ("Usar padrão", f"=COUNTIF('{SHEET}'!E:E,\"Usar padrão\")"),
        ("Configuração", f"=COUNTIF('{SHEET}'!E:E,\"Configuração\")"),
        ("Desenvolvimento", f"=COUNTIF('{SHEET}'!E:E,\"Desenvolvimento\")"),
        ("Fora de escopo", f"=COUNTIF('{SHEET}'!E:E,\"Fora de escopo\")"),
        ("Esforço dev (h)", f"=SUMIF('{SHEET}'!E:E,\"Desenvolvimento\",'{SHEET}'!G:G)"),
    ]
    for i, (k, f) in enumerate(metr):
        r = 4 + i
        a = ws.cell(row=r, column=1, value=k); a.font = C.HEADER_FONT; a.fill = C.HEADER_FILL; a.alignment = C.WRAP
        ws.cell(row=r, column=2, value=f).alignment = C.CENTER
    r = 4 + len(metr) + 1
    note = ws.cell(row=r, column=1,
                   value="Quanto menos desenvolvimento, menor o custo e o risco. "
                         "Reavaliar todo Gap antes de aprovar customização.")
    note.font = C.SUB_FONT; note.alignment = C.WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)


def main(cliente_path="data/exemplo_cliente.yaml", fg_path="data/fitgap.yaml"):
    cli = C.load_yaml(os.path.basename(cliente_path))
    fg = C.load_yaml(os.path.basename(fg_path))
    cliente = cli.get("cliente", {})

    wb = Workbook()
    build_capa(wb, cliente, fg.get("governanca", ""))
    n = build_log(wb, fg.get("itens", []))
    build_resumo(wb, n)

    C.ensure_out()
    fname = f"Log_FitGap_{C.slug(cliente.get('nome'))}.xlsx"
    path = os.path.join(C.OUT, fname)
    wb.save(path)
    print(f"OK: {fname} ({n} itens) -> {path}")


if __name__ == "__main__":
    args = sys.argv[1:]
    main(*args) if args else main()
