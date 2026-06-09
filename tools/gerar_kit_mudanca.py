# -*- coding: utf-8 -*-
"""Gera o Kit de Gestão da Mudança (OCM) em .xlsx, com:
Mapa de Stakeholders, Plano de Comunicação, Prontidão (ADKAR),
Plano de Treinamento por Papel e Indicadores de Adoção.

Uso:
    python gerar_kit_mudanca.py [data/exemplo_cliente.yaml] [data/gestao_mudanca.yaml]
"""
import os
import sys
from openpyxl import Workbook
import _common as C


def build_capa(wb, cliente):
    ws = wb.active
    ws.title = "Capa"
    C.set_widths(ws, [24, 60])
    C.title_block(ws, "Kit de Gestão da Mudança (OCM)",
                  f"{cliente.get('nome','')} · gerado em {C.today()}", span=2)
    info = [
        ("Cliente", cliente.get("nome", "")),
        ("Código SICLA", cliente.get("codigo_sicla", "")),
        ("Usuário líder", cliente.get("usuario_lider", "")),
        ("Virada prevista", cliente.get("data_virada_prevista", "")),
        ("", ""),
        ("Propósito", "Garantir adoção plena: tratar pessoas, comunicação e capacitação, não só o sistema."),
        ("Modelo", "ADKAR — Consciência, Desejo, Conhecimento, Habilidade, Reforço."),
    ]
    for r, (k, v) in enumerate(info, start=4):
        a = ws.cell(row=r, column=1, value=k)
        if k:
            a.font = C.HEADER_FONT; a.fill = C.HEADER_FILL
        a.alignment = C.WRAP
        ws.cell(row=r, column=2, value=v).alignment = C.WRAP


def build_stakeholders(wb, data):
    ws = wb.create_sheet("Mapa de Stakeholders")
    cols = ["Nome/Cargo", "Papel", "Influência", "Interesse", "Postura", "Estratégia de engajamento"]
    C.header_row(ws, cols)
    C.set_widths(ws, [24, 24, 12, 12, 16, 46])
    rows = [[s.get("nome",""), s.get("papel",""), s.get("influencia",""),
             s.get("interesse",""), s.get("postura",""), s.get("estrategia","")]
            for s in data.get("stakeholders", [])]
    C.write_rows(ws, rows)


def build_comunicacao(wb, data):
    ws = wb.create_sheet("Plano de Comunicação")
    cols = ["Momento", "Público", "Canal", "Mensagem-chave", "Responsável", "Frequência"]
    C.header_row(ws, cols)
    C.set_widths(ws, [20, 24, 18, 44, 22, 16])
    rows = [[c.get("momento",""), c.get("publico",""), c.get("canal",""),
             c.get("mensagem",""), c.get("responsavel",""), c.get("frequencia","")]
            for c in data.get("comunicacao", [])]
    C.write_rows(ws, rows)


def build_prontidao(wb, data):
    ws = wb.create_sheet("Prontidão (ADKAR)")
    pront = data.get("prontidao", {})
    dims = pront.get("dimensoes", [])
    grupos = pront.get("grupos", [])
    cols = ["Grupo / Área"] + dims + ["Ações de reforço"]
    C.header_row(ws, cols)
    C.set_widths(ws, [22] + [20] * len(dims) + [34])
    rows = [[g] + [""] * len(dims) + [""] for g in grupos]
    C.write_rows(ws, rows)
    nota = ws.cell(row=len(grupos) + 3, column=1,
                   value="Preencher cada dimensão com nota de 1 (baixo) a 5 (alto) por grupo. Notas baixas viram ações de reforço.")
    nota.font = C.SUB_FONT; nota.alignment = C.WRAP


def build_treinamento(wb, data):
    ws = wb.create_sheet("Treinamento por Papel")
    cols = ["Papel", "Módulos", "Cenários (não telas)", "Carga horária", "Status"]
    C.header_row(ws, cols)
    C.set_widths(ws, [22, 26, 44, 14, 16])
    rows = [[t.get("papel",""), ", ".join(t.get("modulos",[])),
             "\n".join(f"• {c}" for c in t.get("cenarios",[])),
             t.get("carga_horaria",""), "A treinar"]
            for t in data.get("treinamento_papel", [])]
    C.write_rows(ws, rows)


def build_indicadores(wb, data):
    ws = wb.create_sheet("Indicadores de Adoção")
    cols = ["Indicador", "Meta", "Como medir", "Resultado"]
    C.header_row(ws, cols)
    C.set_widths(ws, [40, 26, 36, 18])
    rows = [[i.get("indicador",""), i.get("meta",""), "", ""]
            for i in data.get("indicadores_adocao", [])]
    C.write_rows(ws, rows)


def main(cliente_path="data/exemplo_cliente.yaml", ocm_path="data/gestao_mudanca.yaml"):
    cli = C.load_yaml(os.path.basename(cliente_path))
    ocm = C.load_yaml(os.path.basename(ocm_path))
    cliente = cli.get("cliente", {})

    wb = Workbook()
    build_capa(wb, cliente)
    build_stakeholders(wb, ocm)
    build_comunicacao(wb, ocm)
    build_prontidao(wb, ocm)
    build_treinamento(wb, ocm)
    build_indicadores(wb, ocm)

    C.ensure_out()
    fname = f"Kit_Gestao_Mudanca_{C.slug(cliente.get('nome'))}.xlsx"
    path = os.path.join(C.OUT, fname)
    wb.save(path)
    print(f"OK: {fname} -> {path}")


if __name__ == "__main__":
    args = sys.argv[1:]
    main(*args) if args else main()
