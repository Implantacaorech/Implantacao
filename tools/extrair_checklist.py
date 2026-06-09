# -*- coding: utf-8 -*-
"""
extrair_checklist.py — gera tools/data/checklist_modulos.yaml a partir das abas
de checklist da planilha "Roteiro e Check List por módulos".

Cada linha: Módulo | Adicional | Tipo | Integrações | Item de Go-Live | Menu |
Item | Ação/Observação | Sequência. Vira o "Guia do Consultor" no Levantamento.

Reexecute quando a planilha mudar:
    python extrair_checklist.py ["caminho/Roteiro e Check List ....xlsx"]
"""
import os
import sys
import glob
import yaml
from openpyxl import load_workbook

import _common as C

COLS = ["modulo", "adicional", "tipo", "integracoes", "golive",
        "menu", "item", "acao", "seq"]


def main(path=None):
    if not path:
        cands = glob.glob(os.path.expanduser("~/Downloads/Roteiro e Check List*.xlsx"))
        if not cands:
            sys.exit("Informe o caminho da planilha 'Roteiro e Check List ....xlsx'.")
        path = cands[0]
    wb = load_workbook(path, data_only=True)
    linhas = []
    for sn in wb.sheetnames:
        if sn == "Lista de Módulos":
            continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):       # linha 1 = cabeçalho
            vals = [ws.cell(r, c).value for c in range(1, 10)]
            if not any(v not in (None, "") for v in vals):
                continue
            row = {COLS[i]: ("" if vals[i] is None else str(vals[i]).strip()) for i in range(9)}
            if row["adicional"] and (row["item"] or row["acao"] or row["menu"]):
                linhas.append(row)

    out = os.path.join(C.DATA, "checklist_modulos.yaml")
    with open(out, "w", encoding="utf-8") as f:
        f.write("# Roteiro / Check List por módulo (extraído das abas da planilha).\n")
        f.write("# LOCAL — não versionado; reexecute extrair_checklist.py se a planilha mudar.\n")
        yaml.safe_dump({"linhas": linhas}, f, allow_unicode=True, sort_keys=False, width=200)

    from collections import Counter
    cont = Counter(l["adicional"] for l in linhas)
    print("OK: %d linhas de checklist -> %s" % (len(linhas), out))
    print("   módulos (adicional) com checklist: %d" % len(cont))


if __name__ == "__main__":
    main(*sys.argv[1:])
