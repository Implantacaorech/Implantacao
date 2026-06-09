# -*- coding: utf-8 -*-
"""Gera o Painel de Hypercare (.xlsx): janela de estabilização pós-virada, com
registro de chamados, acompanhamento diário e critérios de saída (gate da
transição para o Suporte).

Uso:
    python gerar_painel_hypercare.py [data/exemplo_cliente.yaml] [data/hypercare.yaml]
"""
import os
import sys
import datetime
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import _common as C


def parse_date(s):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(str(s), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def build_capa(wb, cliente, hc, inicio, fim):
    ws = wb.active
    ws.title = "Capa"
    C.set_widths(ws, [26, 62])
    C.title_block(ws, "Painel de Hypercare", f"{cliente.get('nome','')} · gerado em {C.today()}", span=2)
    info = [
        ("Cliente", cliente.get("nome", "")),
        ("Janela", f"{hc.get('janela_semanas', 4)} semanas"),
        ("Início (virada)", inicio.strftime("%d/%m/%Y") if inicio else "(definir)"),
        ("Fim previsto", fim.strftime("%d/%m/%Y") if fim else "(definir)"),
    ]
    r = 4
    for k, v in info:
        a = ws.cell(row=r, column=1, value=k); a.font = C.HEADER_FONT; a.fill = C.HEADER_FILL; a.alignment = C.WRAP
        ws.cell(row=r, column=2, value=v).alignment = C.WRAP
        r += 1
    ws.cell(row=r + 1, column=1, value="Governança").font = C.TITLE_FONT
    for i, g in enumerate(hc.get("governanca", [])):
        ws.cell(row=r + 2 + i, column=1, value=f"• {g}").alignment = C.WRAP
        ws.merge_cells(start_row=r + 2 + i, start_column=1, end_row=r + 2 + i, end_column=2)


def build_chamados(wb):
    ws = wb.create_sheet("Registro de Chamados")
    cols = ["Data", "Usuário", "Módulo", "Descrição", "Severidade", "Status", "Resolução", "Tempo (h)"]
    C.header_row(ws, cols)
    C.set_widths(ws, [14, 20, 18, 44, 14, 14, 40, 10])
    dv_sev = DataValidation(type="list", formula1='"Crítica,Alta,Média,Baixa"', allow_blank=True)
    dv_st = DataValidation(type="list", formula1='"Aberto,Em andamento,Resolvido"', allow_blank=True)
    ws.add_data_validation(dv_sev); ws.add_data_validation(dv_st)
    dv_sev.add("E2:E300"); dv_st.add("F2:F300")


def build_diario(wb, inicio, dias):
    ws = wb.create_sheet("Acompanhamento Diário")
    cols = ["Dia", "Data", "Chamados abertos", "Resolvidos", "Críticos em aberto", "Adesão %", "Observações"]
    C.header_row(ws, cols)
    C.set_widths(ws, [8, 14, 16, 12, 16, 12, 44])
    rows = []
    for d in range(dias):
        data = (inicio + datetime.timedelta(days=d)).strftime("%d/%m/%Y") if inicio else ""
        rows.append([d + 1, data, "", "", "", "", ""])
    C.write_rows(ws, rows)


def build_criterios(wb, criterios):
    ws = wb.create_sheet("Critérios de Saída")
    cols = ["Critério de saída", "Atingido?", "Evidência"]
    C.header_row(ws, cols)
    C.set_widths(ws, [56, 14, 40])
    rows = [[c, "", ""] for c in criterios]
    C.write_rows(ws, rows)
    dv = DataValidation(type="list", formula1='"Sim,Não,Parcial"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"B2:B{len(rows) + 1}")
    r = len(rows) + 3
    g = ws.cell(row=r, column=1,
                value="GATE: só encerrar o hypercare e transferir ao Suporte com TODOS os critérios atingidos.")
    g.font = C.HEADER_FONT; g.fill = C.HEADER_FILL; g.alignment = C.WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)


def main(cliente_path="data/exemplo_cliente.yaml", hc_path="data/hypercare.yaml"):
    cli = C.load_yaml(os.path.basename(cliente_path))
    hc = C.load_yaml(os.path.basename(hc_path))
    cliente = cli.get("cliente", {})

    semanas = int(hc.get("janela_semanas", 4))
    dias = semanas * 7
    inicio = parse_date(cliente.get("data_virada_prevista"))
    fim = (inicio + datetime.timedelta(days=dias)) if inicio else None

    wb = Workbook()
    build_capa(wb, cliente, hc, inicio, fim)
    build_chamados(wb)
    build_diario(wb, inicio, dias)
    build_criterios(wb, hc.get("criterios_saida", []))

    C.ensure_out()
    fname = f"Painel_Hypercare_{C.slug(cliente.get('nome'))}.xlsx"
    path = os.path.join(C.OUT, fname)
    wb.save(path)
    print(f"OK: {fname} (janela {semanas} semanas / {dias} dias) -> {path}")


if __name__ == "__main__":
    args = sys.argv[1:]
    main(*args) if args else main()
