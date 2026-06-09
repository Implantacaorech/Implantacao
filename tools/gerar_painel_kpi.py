# -*- coding: utf-8 -*-
"""Gera o Painel de KPIs da implantação (.xlsx): indicadores de resultado,
marcos (prazo) e horas (planejado x real).

Uso:
    python gerar_painel_kpi.py [data/exemplo_cliente.yaml] [data/kpi.yaml]
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import _common as C


def build_capa(wb, cliente):
    ws = wb.active
    ws.title = "Capa"
    C.set_widths(ws, [26, 60])
    C.title_block(ws, "Painel de KPIs da Implantação",
                  f"{cliente.get('nome','')} · gerado em {C.today()}", span=2)
    for r, (k, v) in enumerate([
        ("Cliente", cliente.get("nome", "")),
        ("Código SICLA", cliente.get("codigo_sicla", "")),
        ("RNS de Implantação", cliente.get("rns_implantacao", "")),
        ("Virada prevista", cliente.get("data_virada_prevista", "")),
    ], start=4):
        a = ws.cell(row=r, column=1, value=k); a.font = C.HEADER_FONT; a.fill = C.HEADER_FILL; a.alignment = C.WRAP
        ws.cell(row=r, column=2, value=v).alignment = C.WRAP


def build_kpis(wb, indicadores):
    ws = wb.create_sheet("Painel de KPIs")
    cols = ["KPI", "Categoria", "Meta", "Medição", "Valor atual", "Farol"]
    C.header_row(ws, cols)
    C.set_widths(ws, [30, 16, 30, 34, 16, 12])
    rows = [[i.get("kpi",""), i.get("categoria",""), i.get("meta",""),
             i.get("medicao",""), "", ""] for i in indicadores]
    C.write_rows(ws, rows)
    dv = DataValidation(type="list", formula1='"Verde,Amarelo,Vermelho"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F2:F{len(rows) + 1}")


def build_marcos(wb, marcos):
    ws = wb.create_sheet("Marcos (Prazo)")
    cols = ["Marco", "Data prevista", "Data real", "Desvio (dias)"]
    C.header_row(ws, cols)
    C.set_widths(ws, [34, 16, 16, 14])
    for i, m in enumerate(marcos):
        r = i + 2
        ws.cell(row=r, column=1, value=m).alignment = C.WRAP
        ws.cell(row=r, column=4, value=f"=IF(AND(B{r}<>\"\",C{r}<>\"\"),C{r}-B{r},\"\")").alignment = C.CENTER
        for col in (2, 3):
            ws.cell(row=r, column=col).border = C.BORDER


def build_horas(wb, horas):
    ws = wb.create_sheet("Horas (Plan x Real)")
    cols = ["Macro-etapa", "Horas planejadas", "Horas reais", "Desvio", "% do plano"]
    C.header_row(ws, cols)
    C.set_widths(ws, [30, 16, 16, 12, 12])
    for i, h in enumerate(horas):
        r = i + 2
        ws.cell(row=r, column=1, value=h.get("etapa", "")).alignment = C.WRAP
        ws.cell(row=r, column=4, value=f"=IF(AND(B{r}<>\"\",C{r}<>\"\"),C{r}-B{r},\"\")").alignment = C.CENTER
        ws.cell(row=r, column=5, value=f"=IF(B{r}>0,C{r}/B{r},\"\")").alignment = C.CENTER
        ws.cell(row=r, column=5).number_format = "0%"
        for col in (2, 3):
            ws.cell(row=r, column=col).border = C.BORDER
    # Linha de total
    r = len(horas) + 2
    t = ws.cell(row=r, column=1, value="TOTAL"); t.font = C.HEADER_FONT
    ws.cell(row=r, column=2, value=f"=SUM(B2:B{r-1})")
    ws.cell(row=r, column=3, value=f"=SUM(C2:C{r-1})")


def main(cliente_path="data/exemplo_cliente.yaml", kpi_path="data/kpi.yaml"):
    cli = C.load_yaml(os.path.basename(cliente_path))
    kpi = C.load_yaml(os.path.basename(kpi_path))
    cliente = cli.get("cliente", {})

    wb = Workbook()
    build_capa(wb, cliente)
    build_kpis(wb, kpi.get("indicadores", []))
    build_marcos(wb, kpi.get("marcos", []))
    build_horas(wb, kpi.get("horas", []))

    C.ensure_out()
    fname = f"Painel_KPIs_{C.slug(cliente.get('nome'))}.xlsx"
    path = os.path.join(C.OUT, fname)
    wb.save(path)
    print(f"OK: {fname} ({len(kpi.get('indicadores', []))} KPIs) -> {path}")


if __name__ == "__main__":
    args = sys.argv[1:]
    main(*args) if args else main()
