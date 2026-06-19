# -*- coding: utf-8 -*-
"""
extrair_indice_topicos.py — gera tools/data/indice_topicos.yaml a partir da
planilha "Indice de Topicos para Mapeamento de Processos.xlsx".

A planilha é hierárquica (uma única aba):
    Nº | Sigla | Módulos | Adicionais | Tópicos
  - linha de Módulo principal: Nº + Sigla + Módulos (col C)
  - linha de Adicional:        Nº + Sigla + Adicionais (col D)  -> pertence ao módulo acima
  - linha de Tópico:           Tópicos (col E)                  -> pertence ao módulo/adicional acima

Cada tópico vira uma linha achatada com TODAS as colunas (contexto do módulo e
do adicional), preservando a ordem da planilha. Usado pelo Cadastro do Índice.

Reexecute quando a planilha mudar:
    python extrair_indice_topicos.py ["caminho/Indice de Topicos ....xlsx"]
"""
import os
import sys
import glob
import yaml
from openpyxl import load_workbook

import _common as C

COLS = ["modulo_num", "modulo_sigla", "modulo",
        "adicional_num", "adicional_sigla", "adicional", "topico"]


def _s(v):
    return "" if v is None else str(v).strip()


def extrair(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = []
    m_num = m_sig = m_nome = ""
    a_num = a_sig = a_nome = ""
    ordem = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue  # cabeçalho
        num, sig = _s(row[0]), _s(row[1])
        modulo, adicional, topico = _s(row[2]), _s(row[3]), _s(row[4])
        if modulo:                     # novo Módulo principal
            m_num, m_sig, m_nome = num, sig, modulo
            a_num = a_sig = a_nome = ""
        elif adicional:                # novo Adicional (sob o módulo atual)
            a_num, a_sig, a_nome = num, sig, adicional
        elif topico:                   # Tópico (sob o módulo/adicional atual)
            ordem += 1
            linhas.append({"modulo_num": m_num, "modulo_sigla": m_sig, "modulo": m_nome,
                           "adicional_num": a_num, "adicional_sigla": a_sig, "adicional": a_nome,
                           "topico": topico, "ordem": ordem})
    return linhas


def main(path=None):
    if not path:
        cands = glob.glob(os.path.expanduser("~/Downloads/Indice de Topicos*.xlsx"))
        if not cands:
            sys.exit("Informe o caminho da planilha 'Indice de Topicos ....xlsx'.")
        path = cands[0]
    linhas = extrair(path)
    out = os.path.join(C.DATA, "indice_topicos.yaml")
    with open(out, "w", encoding="utf-8") as f:
        f.write("# Índice de Tópicos para Mapeamento de Processos (extraído da planilha).\n")
        f.write("# Reexecute extrair_indice_topicos.py se a planilha mudar.\n")
        yaml.safe_dump({"linhas": linhas}, f, allow_unicode=True, sort_keys=False, width=200)

    mods = {(l["modulo_sigla"], l["modulo"]) for l in linhas}
    adic = {(l["adicional_sigla"], l["adicional"]) for l in linhas if l["adicional"]}
    print("OK: %d tópicos -> %s" % (len(linhas), out))
    print("   módulos: %d | adicionais: %d" % (len(mods), len(adic)))


if __name__ == "__main__":
    main(*sys.argv[1:])
