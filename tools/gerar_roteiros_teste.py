# -*- coding: utf-8 -*-
"""Gera a planilha de Roteiros SIT/UAT (.xlsx) por módulo, com registro de
defeitos e painel de sign-off (gate da virada).

Uso:
    python gerar_roteiros_teste.py [data/exemplo_cliente.yaml] [data/roteiros_teste.yaml]
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import _common as C

STATUS_OPCOES = '"Não testado,Aprovado,Reprovado,Bloqueado"'
COLS = ["ID", "Tipo", "Criticidade", "Cenário", "Pré-requisitos", "Passos",
        "Resultado esperado", "Status", "Resultado obtido", "Defeito", "Testado por", "Data"]
WIDTHS = [10, 8, 12, 34, 28, 40, 38, 14, 30, 12, 16, 12]


def add_modulo_sheet(wb, mod):
    ws = wb.create_sheet(C.safe_sheet(mod["nome"]))
    C.header_row(ws, COLS)
    C.set_widths(ws, WIDTHS)
    rows = []
    for caso in mod.get("casos", []):
        passos = "\n".join(f"{i}. {p}" for i, p in enumerate(caso.get("passos", []), 1))
        rows.append([
            caso.get("id", ""), caso.get("tipo", ""), caso.get("criticidade", ""),
            caso.get("cenario", ""), caso.get("pre_requisitos", ""), passos,
            caso.get("resultado_esperado", ""), "Não testado", "", "", "", "",
        ])
    C.write_rows(ws, rows)
    n = len(rows)
    if n:
        dv = DataValidation(type="list", formula1=STATUS_OPCOES, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"H2:H{n + 1}")
    return ws.title, n


def build_capa(wb, cliente, modulos):
    ws = wb.active
    ws.title = "Capa"
    C.set_widths(ws, [24, 60])
    C.title_block(ws, "Roteiros de Teste — SIT / UAT",
                  f"{cliente.get('nome','')} · gerado em {C.today()}", span=2)
    info = [
        ("Cliente", cliente.get("nome", "")),
        ("Código SICLA", cliente.get("codigo_sicla", "")),
        ("RNS de Implantação", cliente.get("rns_implantacao", "")),
        ("Virada prevista", cliente.get("data_virada_prevista", "")),
        ("Módulos no escopo", ", ".join(modulos)),
        ("", ""),
        ("SIT", "Teste Integrado — executado pelo consultor (integração entre módulos)"),
        ("UAT", "Aceite do Usuário — executado pelo cliente (valida o processo real)"),
        ("Legenda Status", "Não testado · Aprovado · Reprovado · Bloqueado"),
        ("Gate da virada", "≥ 95% dos casos UAT Aprovados e 0 defeitos de severidade Crítica"),
    ]
    for r, (k, v) in enumerate(info, start=4):
        a = ws.cell(row=r, column=1, value=k); a.font = C.HEADER_FONT if k else C.SUB_FONT
        if k:
            a.fill = C.HEADER_FILL; a.alignment = C.WRAP
        b = ws.cell(row=r, column=2, value=v); b.alignment = C.WRAP


def build_defeitos(wb):
    ws = wb.create_sheet("Registro de Defeitos")
    cols = ["ID", "Caso (ID)", "Módulo", "Descrição do defeito", "Severidade",
            "Status", "Responsável", "Aberto em", "Resolvido em"]
    C.header_row(ws, cols)
    C.set_widths(ws, [10, 12, 22, 46, 14, 14, 18, 14, 14])
    dv_sev = DataValidation(type="list", formula1='"Crítica,Alta,Média,Baixa"', allow_blank=True)
    dv_st = DataValidation(type="list", formula1='"Aberto,Em análise,Resolvido,Fechado"', allow_blank=True)
    ws.add_data_validation(dv_sev); ws.add_data_validation(dv_st)
    dv_sev.add("E2:E200"); dv_st.add("F2:F200")


def build_resumo(wb, sheet_names):
    ws = wb.create_sheet("Resumo e Sign-off")
    C.set_widths(ws, [28, 18, 50])
    C.title_block(ws, "Resumo dos Testes e Liberação", span=3)
    # Métricas vivas (COUNTIF sobre a coluna H de cada aba de módulo)
    def soma(status):
        partes = [f"COUNTIF('{s}'!H:H,\"{status}\")" for s in sheet_names]
        return "=" + "+".join(partes) if partes else "=0"
    metr = [
        ("Casos Aprovados", soma("Aprovado")),
        ("Casos Reprovados", soma("Reprovado")),
        ("Casos Bloqueados", soma("Bloqueado")),
        ("Casos Não testados", soma("Não testado")),
    ]
    r0 = 4
    for i, (k, formula) in enumerate(metr):
        r = r0 + i
        a = ws.cell(row=r, column=1, value=k); a.font = C.HEADER_FONT; a.fill = C.HEADER_FILL; a.alignment = C.WRAP
        ws.cell(row=r, column=2, value=formula).alignment = C.CENTER
    # Critério de liberação
    rg = r0 + len(metr) + 1
    g = ws.cell(row=rg, column=1, value="Gate da virada")
    g.font = C.HEADER_FONT; g.fill = C.HEADER_FILL; g.alignment = C.WRAP
    ws.merge_cells(start_row=rg, start_column=2, end_row=rg, end_column=3)
    ws.cell(row=rg, column=2,
            value="Liberar a virada somente com ≥ 95% dos casos UAT Aprovados e 0 defeitos Críticos em aberto."
            ).alignment = C.WRAP
    # Assinaturas
    rs = rg + 2
    ws.cell(row=rs, column=1, value="Assinaturas").font = C.TITLE_FONT
    for i, papel in enumerate(["Consultor de Implantação", "Usuário Líder (cliente)", "Gerente do Projeto"]):
        r = rs + 2 + i * 2
        ws.cell(row=r, column=1, value="__________________________")
        ws.cell(row=r + 1, column=1, value=papel).font = C.SUB_FONT


def main(cliente_path="data/exemplo_cliente.yaml", roteiros_path="data/roteiros_teste.yaml"):
    cli = C.load_yaml(os.path.basename(cliente_path))
    rot = C.load_yaml(os.path.basename(roteiros_path))
    cliente = cli.get("cliente", {})
    modulos = rot.get("modulos", [])

    wb = Workbook()
    build_capa(wb, cliente, [m["nome"] for m in modulos])
    sheet_names = []
    total = 0
    for mod in modulos:
        name, n = add_modulo_sheet(wb, mod)
        sheet_names.append(name)
        total += n
    build_defeitos(wb)
    build_resumo(wb, sheet_names)

    C.ensure_out()
    fname = f"Roteiros_SIT_UAT_{C.slug(cliente.get('nome'))}.xlsx"
    path = os.path.join(C.OUT, fname)
    wb.save(path)
    print(f"OK: {fname} ({total} casos em {len(sheet_names)} módulos) -> {path}")


if __name__ == "__main__":
    args = sys.argv[1:]
    main(*args) if args else main()
