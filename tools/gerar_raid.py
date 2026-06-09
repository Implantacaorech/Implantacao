# -*- coding: utf-8 -*-
"""Gera o RAID (.xlsx) — Riscos, Premissas (Assumptions), Issues, Decisões e
Dependências por projeto de implantação.

Uso:
    python gerar_raid.py [data/exemplo_cliente.yaml] [data/raid.yaml]
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import _common as C


def _sheet(wb, nome, cols, widths, rows, dropdowns=None):
    ws = wb.create_sheet(nome)
    C.header_row(ws, cols)
    C.set_widths(ws, widths)
    C.write_rows(ws, rows)
    n = max(len(rows), 1)
    for col_letter, opcoes in (dropdowns or {}).items():
        dv = DataValidation(type="list", formula1=f'"{opcoes}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{n + 50}")
    return ws


def build_capa(wb, cliente):
    ws = wb.active
    ws.title = "Capa"
    C.set_widths(ws, [26, 60])
    C.title_block(ws, "RAID — Riscos, Premissas, Issues e Decisões",
                  f"{cliente.get('nome','')} · gerado em {C.today()}", span=2)
    legenda = [
        ("R", "Riscos — eventos futuros que podem impactar o projeto"),
        ("A", "Premissas (Assumptions) — o que assumimos como verdade"),
        ("I", "Issues — problemas que já ocorreram e precisam de ação"),
        ("D", "Decisões e Dependências — o que foi decidido / o que depende de terceiros"),
    ]
    for r, (k, v) in enumerate(legenda, start=4):
        a = ws.cell(row=r, column=1, value=k); a.font = C.HEADER_FONT; a.fill = C.HEADER_FILL; a.alignment = C.CENTER
        ws.cell(row=r, column=2, value=v).alignment = C.WRAP


def main(cliente_path="data/exemplo_cliente.yaml", raid_path="data/raid.yaml"):
    cli = C.load_yaml(os.path.basename(cliente_path))
    raid = C.load_yaml(os.path.basename(raid_path))
    cliente = cli.get("cliente", {})

    wb = Workbook()
    build_capa(wb, cliente)

    # Riscos
    _sheet(wb, "Riscos",
           ["ID", "Descrição", "Impacto", "Probabilidade", "Mitigação", "Responsável", "Status"],
           [8, 46, 12, 14, 44, 18, 14],
           [[f"R-{i:02d}", r.get("descricao",""), r.get("impacto",""), r.get("probabilidade",""),
             r.get("mitigacao",""), r.get("responsavel",""), "Aberto"]
            for i, r in enumerate(raid.get("riscos", []), 1)],
           {"C": "Alto,Médio,Baixo", "D": "Alta,Média,Baixa", "G": "Aberto,Mitigado,Fechado"})

    # Premissas
    _sheet(wb, "Premissas",
           ["ID", "Descrição", "Validada?", "Observação"],
           [8, 56, 12, 36],
           [[f"A-{i:02d}", p.get("descricao",""), p.get("validada",""), ""]
            for i, p in enumerate(raid.get("premissas", []), 1)],
           {"C": "Sim,Não,Parcial"})

    # Issues
    _sheet(wb, "Issues",
           ["ID", "Descrição", "Severidade", "Ação", "Responsável", "Status"],
           [8, 46, 14, 40, 18, 14],
           [[f"I-{i:02d}", s.get("descricao",""), s.get("severidade",""), s.get("acao",""),
             s.get("responsavel",""), s.get("status","Aberto")]
            for i, s in enumerate(raid.get("issues", []), 1)],
           {"C": "Crítica,Alta,Média,Baixa", "F": "Aberto,Em ação,Resolvido"})

    # Decisões
    _sheet(wb, "Decisões",
           ["ID", "Descrição", "Data", "Decidido por"],
           [8, 56, 14, 24],
           [[f"D-{i:02d}", d.get("descricao",""), d.get("data",""), d.get("por","")]
            for i, d in enumerate(raid.get("decisoes", []), 1)])

    # Dependências
    _sheet(wb, "Dependências",
           ["ID", "Descrição", "De quem", "Prazo", "Status"],
           [8, 50, 24, 14, 16],
           [[f"DP-{i:02d}", dp.get("descricao",""), dp.get("de_quem",""), dp.get("prazo",""),
             dp.get("status","Em andamento")]
            for i, dp in enumerate(raid.get("dependencias", []), 1)],
           {"E": "Não iniciado,Em andamento,Concluído,Bloqueado"})

    C.ensure_out()
    fname = f"RAID_{C.slug(cliente.get('nome'))}.xlsx"
    path = os.path.join(C.OUT, fname)
    wb.save(path)
    print(f"OK: {fname} -> {path}")


if __name__ == "__main__":
    args = sys.argv[1:]
    main(*args) if args else main()
