# -*- coding: utf-8 -*-
"""Importador da Matriz de Conhecimento (docs/Matriz de Conhecimento.xlsx).

Estrutura da planilha (aba 'Matriz'): linha 7 = áreas (células mescladas),
linha 8 = siglas das competências (+ colunas fixas Nome/Dias/Setor/Ár),
linhas 9+ = um técnico por linha com notas 1-10.

Regra de importação: cria competências e técnicos que NÃO existem; técnico
já existente NÃO é sobrescrito (preserva as edições feitas na tela).
Uso:  python webapp/matriz.py   (ou botão Importar na tela, só Administrador)
"""
import os
import json
import logging
from datetime import datetime

import db

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_PADRAO = os.path.join(os.path.dirname(HERE), "docs", "Matriz de Conhecimento.xlsx")
FIXAS = {"Nome", "Dias", "Setor", "Ár", "Dt.última atu"}


def _limpa_area(a):
    a = (a or "").replace("AREA DE CONHECIMENTO -->", "").strip()
    troca = {"AREA: C": "Controladoria", "AREA: R": "Folha de Pagamento", "ÁREA N": "Negócios",
             "AREA F": "Finanças", "AREA: P": "Produção", "GERAIS": "Gerais",
             "OUTRAS ROTINAS/COMPETÊNCIAS RELEVANTES": "Outras rotinas", "FORMULÁRIOS": "Formulários"}
    return troca.get(a, a or "Outras rotinas")


def ler_planilha(path=None):
    """Lê a planilha e devolve (competencias [(sigla, area, ordem)], tecnicos [dict])."""
    from openpyxl import load_workbook
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(path or XLSX_PADRAO, data_only=True)
    ws = wb["Matriz"] if "Matriz" in wb.sheetnames else wb[wb.sheetnames[0]]

    area_por_col, atual = {}, ""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=7, column=c).value
        if v and str(v).strip():
            atual = str(v).strip()
        area_por_col[c] = atual

    comps, col_sigla, col_nome, col_dias, col_setor = [], {}, 2, 3, 4
    ordem = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=8, column=c).value
        s = str(v).strip() if v not in (None, "") else ""
        if not s:
            continue
        if s == "Nome":
            col_nome = c
        elif s == "Dias":
            col_dias = c
        elif s == "Setor":
            col_setor = c
        if s in FIXAS:
            continue
        ordem += 1
        comps.append((s, _limpa_area(area_por_col[c]), ordem))
        col_sigla[c] = s

    tecnicos = []
    for r in range(9, ws.max_row + 1):
        nome = ws.cell(row=r, column=col_nome).value
        if not (nome and str(nome).strip()):
            continue
        notas = {}
        for c, sigla in col_sigla.items():
            v = ws.cell(row=r, column=c).value
            if v in (None, ""):
                continue
            try:
                n = int(float(str(v).replace(",", ".")))
            except (TypeError, ValueError):
                continue
            notas[sigla] = max(0, min(10, n))
        tecnicos.append({"nome": str(nome).strip(),
                         "setor": str(ws.cell(row=r, column=col_setor).value or "").strip(),
                         "dias": str(ws.cell(row=r, column=col_dias).value or "").strip(),
                         "notas": notas})
    return comps, tecnicos


def importar(path=None, autor="importação"):
    """Importa para o banco (aditivo — não sobrescreve técnico existente).
    Devolve (novas_competencias, novos_tecnicos, ignorados)."""
    comps, tecnicos = ler_planilha(path)
    novas = novos = ignorados = 0
    with db.Session() as s:
        existentes = {c.sigla for c in s.query(db.MatrizCompetencia).all()}
        for sigla, area, ordem in comps:
            if sigla in existentes:
                continue
            s.add(db.MatrizCompetencia(sigla=sigla, area=area, ordem=ordem))
            novas += 1
        nomes = {(t.nome or "").strip().lower() for t in s.query(db.MatrizTecnico).all()}
        for t in tecnicos:
            if t["nome"].lower() in nomes:
                ignorados += 1
                continue
            s.add(db.MatrizTecnico(nome=t["nome"], setor=t["setor"], dias=t["dias"],
                                   notas=json.dumps(t["notas"], ensure_ascii=False),
                                   atualizado_em=datetime.now(), atualizado_por=autor))
            novos += 1
        s.commit()
    logging.info("Matriz: %d competências novas, %d técnicos novos, %d preservados.",
                 novas, novos, ignorados)
    return novas, novos, ignorados


if __name__ == "__main__":
    n_c, n_t, ig = importar()
    print("Competências novas: %d | Técnicos novos: %d | Já existiam (preservados): %d" % (n_c, n_t, ig))
