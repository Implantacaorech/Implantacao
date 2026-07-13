# -*- coding: utf-8 -*-
"""Geração fiel — saídas em XLSX (cronograma do layout + cronograma de visitas)."""
import os

import _common as C
import db
from gl_comum import _norm


def _preencher_cronograma_xlsx(wb, projeto):
    """Preenche o cabeçalho (Consultor/Horas) por rótulo e as linhas de visita a partir
    do cronograma do projeto, sem tocar nas colunas com fórmula (Total/Horário)."""
    ws = wb["Cronograma de visitas"] if "Cronograma de visitas" in wb.sheetnames else wb.worksheets[0]
    alvos = {
        "consultor": (projeto.get("consultor") or "").strip(),
        "horas do planejamento": str(projeto.get("horas_cobradas") or "").strip(),
        "hrs previstas bonificadas": str(projeto.get("horas_bonificadas") or "").strip(),
    }
    for row in ws.iter_rows(min_row=1, max_row=8):       # cabeçalho: valor à direita do rótulo
        for c in row:
            if isinstance(c.value, str):
                v = alvos.get(_norm(c.value))
                if v:
                    ws.cell(row=c.row, column=c.column + 1, value=v)
    itens = db.cronograma_do_projeto(projeto.get("id")) if projeto.get("id") else []
    if not itens:
        return 0
    hdr, cols = None, {}
    for row in ws.iter_rows(min_row=1, max_row=15):       # acha a linha de cabeçalho da tabela
        vals = {_norm(c.value): c.column for c in row if isinstance(c.value, str)}
        if "data" in vals and "o que será abordado" in vals:
            hdr, cols = row[0].row, vals
            break
    if not hdr:
        return 0
    cons = (projeto.get("consultor") or "").strip()
    for i, it in enumerate(itens):
        r = hdr + 1 + i

        def setc(label, val):
            col = cols.get(label)
            if col and val:
                ws.cell(row=r, column=col, value=val)

        abordado = (it.get("etapa") or "").strip()
        if it.get("topicos"):
            abordado = (abordado + " — " + it["topicos"]).strip(" —")
        setc("data", it.get("data", ""))
        setc("local", it.get("modalidade", ""))
        setc("técnico", cons)
        setc("o que será abordado", abordado)
    return len(itens)


def _logo_rech():
    """Caminho da logo Rech (webapp/static) — None se o arquivo não existir (não quebra a
    geração; só sai sem a imagem no cabeçalho)."""
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "logo-rech-color.png")
    return p if os.path.exists(p) else None


def gerar_agenda_xlsx(projeto, atividades, horarios=None):
    """Gera o cronograma de visitas (.xlsx) a partir das atividades ALOCADAS (data+turno)
    do agendador. `horarios` = horário global por turno (db.cronograma_horarios). Cabeçalho
    com dados do cliente + logo Rech; cada linha traz também o Analista Responsável do módulo
    (sobreposição em Designacao.analista, senão o padrão do projeto). Devolve o caminho."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import datetime as _dt
    horarios = horarios or {}
    DIAS = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    TURNOS = {"manha": "Manhã", "tarde": "Tarde"}
    aloc = [a for a in atividades if (a.get("data") and a.get("turno"))]
    aloc.sort(key=lambda a: (a["data"], 0 if a["turno"] == "manha" else 1, a["modulo"], a["seq"]))

    pid = projeto.get("id")
    designacoes = db.designacoes_do_projeto(pid) if pid else []
    analista_mod = {d["modulo"]: (d.get("analista") or "").strip() for d in designacoes}
    analista_padrao = ((db.cronograma_config(pid) or {}).get("analista_padrao") or "").strip() if pid else ""

    def analista_de(modulo):
        return analista_mod.get(modulo) or analista_padrao

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma de Visitas"

    # Cabeçalho: dados do cliente (colunas C+; A/B reservadas para a logo) + logo Rech em A1.
    ws.append([])
    ws.append(["", "", "Cronograma de Visitas — %s" % (projeto.get("cliente") or "")])
    ws.append(["", "", "Cliente:", projeto.get("cliente") or "", "", "CNPJ:", projeto.get("cnpj") or ""])
    ws.append(["", "", "Nº Projeto:", projeto.get("numero_projeto") or "",
              "", "Consultor(es):", projeto.get("consultor") or ""])
    ws.append([])
    ws["C2"].font = Font(bold=True, size=13, color="1F4E79")
    for cel in ("C3", "F3", "C4", "F4"):
        ws[cel].font = Font(bold=True, size=9, color="595959")
    for cel in ("D3", "G3", "D4", "G4"):
        ws[cel].font = Font(size=9)
    ws.row_dimensions[1].height = 6
    for r in (2, 3, 4):
        ws.row_dimensions[r].height = 16

    logo = _logo_rech()
    if logo:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo)
            img.width, img.height = 150, 47   # mantém a proporção original (250x79)
            ws.add_image(img, "A1")
        except Exception:
            pass   # ambiente sem Pillow ou logo corrompida — segue sem a imagem, não quebra a geração

    cab = ["Data", "Dia", "Turno", "Horário", "Módulo", "Visita", "Atividade", "Tipo", "Técnico",
           "Analista Responsável", "Status"]
    ws.append(cab)
    azul = PatternFill("solid", fgColor="1F4E79")
    borda = Border(*(Side(style="thin", color="D0D7E2"),) * 4)
    hdr_row = ws.max_row
    for c in ws[hdr_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul
        c.alignment = Alignment(horizontal="center")
    for a in aloc:
        try:
            d = _dt.date.fromisoformat(a["data"])
            data_lbl, dia_lbl = d.strftime("%d/%m/%Y"), DIAS[d.weekday()]
        except ValueError:
            data_lbl, dia_lbl = a["data"], ""
        ini, fim = db.cronograma_horas(horarios, a["turno"])
        hora_lbl = ("%s–%s" % (ini, fim)) if (ini or fim) else ""
        ws.append([data_lbl, dia_lbl, TURNOS.get(a["turno"], a["turno"]), hora_lbl, a["modulo"],
                   "V%s" % a["seq"], a.get("descricao", ""), a.get("tipo", ""),
                   a.get("tecnico", ""), analista_de(a["modulo"]), a.get("status", "")])
    for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row):
        for c in row:
            c.border = borda
            c.alignment = Alignment(vertical="top", wrap_text=True)
    larg = [14, 6, 9, 12, 9, 8, 52, 16, 20, 20, 12]
    for i, w in enumerate(larg):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    ws.freeze_panes = "A%d" % (hdr_row + 1)

    os.makedirs(C.OUT, exist_ok=True)
    dest = os.path.join(C.OUT, "cronograma_visitas_%s.xlsx" % C.slug(projeto.get("cliente") or "cliente"))
    wb.save(dest)
    return dest
