# -*- coding: utf-8 -*-
"""Pré-visualização (WYSIWYG) de documentos gerados: renderiza .docx/.xlsx como HTML
para ver o documento na tela (folha A4), sem precisar baixar e abrir o Office."""
import os
import html
import threading

_PDF_LOCK = threading.Lock()   # serializa as conversões (Word COM não gosta de concorrência)


def to_pdf(path):
    """Caminho de um PDF FIEL do .docx — renderizado pelo próprio Word (espelho exato),
    com cache por arquivo+mtime. Devolve None se não for .docx, se o Word não estiver
    disponível ou se a conversão falhar (o chamador cai no HTML). A conversão roda num
    SUBPROCESSO isolado, para não desestabilizar o servidor web."""
    if os.path.splitext(path)[1].lower() != ".docx" or not os.path.exists(path):
        return None
    import hashlib
    import subprocess
    import sys
    try:
        import _common as C
        base = C.DATA_WRITE
    except Exception:
        base = os.path.dirname(os.path.abspath(path))
    cache = os.path.join(base, "_preview")
    key = hashlib.md5(("%s|%s" % (os.path.abspath(path), os.path.getmtime(path))).encode()).hexdigest()
    pdf = os.path.join(cache, key + ".pdf")
    if os.path.exists(pdf) and os.path.getsize(pdf) > 0:
        return pdf
    with _PDF_LOCK:
        if os.path.exists(pdf) and os.path.getsize(pdf) > 0:
            return pdf
        os.makedirs(cache, exist_ok=True)
        try:
            subprocess.run([sys.executable, os.path.abspath(__file__),
                            os.path.abspath(path), os.path.abspath(pdf)],
                           capture_output=True, timeout=120)
        except Exception:
            import logging
            logging.exception("Falha ao converter docx->pdf (%s)", path)
        if os.path.exists(pdf) and os.path.getsize(pdf) > 0:
            return pdf
    return None


def _convert_docx_to_pdf(src, dst):
    """Conversão Word COM (Windows). Roda em subprocesso isolado — ver o bloco __main__."""
    import pythoncom
    import win32com.client
    pythoncom.CoInitialize()
    word = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        doc = word.Documents.Open(os.path.abspath(src), ReadOnly=True)
        doc.SaveAs(os.path.abspath(dst), FileFormat=17)   # 17 = wdFormatPDF
        doc.Close(False)
    finally:
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _esc(t):
    return html.escape(str(t if t is not None else ""))


def to_html(path):
    """HTML do conteúdo do documento. Suporta .docx e .xlsx; senão, aviso."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".docx":
        return _docx_html(path)
    if ext in (".xlsx", ".xlsm"):
        return _xlsx_html(path)
    return "<p class='aviso'>Pré-visualização indisponível para este tipo de arquivo. Use “Baixar”.</p>"


# ---------- .docx ----------
def _iter_blocks(doc):
    """Itera parágrafos e tabelas na ordem real do documento."""
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph
    body = doc.element.body
    for child in body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, doc)
        elif isinstance(child, CT_Tbl):
            yield Table(child, doc)


def _runs_html(par):
    out = []
    for r in par.runs:
        t = _esc(r.text)
        if not t:
            continue
        if r.bold:
            t = "<strong>%s</strong>" % t
        if r.italic:
            t = "<em>%s</em>" % t
        out.append(t)
    return "".join(out) or _esc(par.text)


def _par_html(par):
    txt = _runs_html(par)
    if not txt.strip():
        return ""
    st = (par.style.name or "").lower() if par.style else ""
    if "title" in st:
        return "<h1>%s</h1>" % txt
    if "heading 1" in st:
        return "<h2>%s</h2>" % txt
    if "heading" in st:
        return "<h3>%s</h3>" % txt
    if "list" in st:
        return "<p class='li'>• %s</p>" % txt
    return "<p>%s</p>" % txt


def _tab_html(tab):
    trs = []
    for ri, row in enumerate(tab.rows):
        tag = "th" if ri == 0 else "td"
        tds = "".join("<%s>%s</%s>" % (tag, _esc(c.text), tag) for c in row.cells)
        trs.append("<tr>%s</tr>" % tds)
    return "<table class='doc-tab'>%s</table>" % "".join(trs)


def _docx_html(path):
    from docx import Document
    from docx.table import Table
    from docx.text.paragraph import Paragraph
    doc = Document(path)
    parts = []
    for blk in _iter_blocks(doc):
        if isinstance(blk, Paragraph):
            parts.append(_par_html(blk))
        elif isinstance(blk, Table):
            parts.append(_tab_html(blk))
    return "\n".join(p for p in parts if p) or "<p class='aviso'>(documento sem conteúdo de texto)</p>"


# ---------- .xlsx ----------
def _xlsx_html(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append("<h2>%s</h2>" % _esc(ws.title))
        rows = [r for r in ws.iter_rows(values_only=True)
                if any((c is not None and str(c).strip()) for c in r)]
        if not rows:
            out.append("<p class='aviso'>(planilha vazia)</p>")
            continue
        trs = []
        for ri, r in enumerate(rows):
            tag = "th" if ri == 0 else "td"
            tds = "".join("<%s>%s</%s>" % (tag, _esc(c), tag) for c in r)
            trs.append("<tr>%s</tr>" % tds)
        out.append("<table class='doc-tab'>%s</table>" % "".join(trs))
    wb.close()
    return "\n".join(out)


if __name__ == "__main__":   # subprocesso de conversão: python docview.py <src.docx> <dst.pdf>
    import sys
    if len(sys.argv) == 3:
        try:
            _convert_docx_to_pdf(sys.argv[1], sys.argv[2])
        except Exception as e:
            sys.stderr.write("%s: %s\n" % (type(e).__name__, e))
            sys.exit(1)
